import datetime
from math import prod
import re
import traceback
import openpyxl as openxl
from openpyxl import load_workbook
from pytils import translit
from django.utils.text import slugify
from simple_history.utils import update_change_reason

from apps import supplier
from apps.core.models import Currency, Vat
from apps.logs.utils import error_alert
from apps.product.models import Lot, Price, Product, Stock
from apps.supplier.models import Supplier, Vendor
from project.settings import MEDIA_ROOT


def get_motrum_storage(path):
    try:
        new_dir = "{0}/{1}".format(MEDIA_ROOT, path)
        path_storage = new_dir
        
        # path_storage_motrum = f"{new_dir}/test_ooo.xlsx"
        # path_storage_pnm = f"{new_dir}/test_pmn.xlsx"

        # def get_group_rows(sheet):
        #     # a =  [row[0].row for row in sheet.iter_rows(2) if row[0].alignment.indent == 0.0]
        #     row_index_groupe = [
        #         row_index
        #         for row_index, row_dimension in sheet.row_dimensions.items()
        #         if row_index > 2 and row_dimension.outline_level == 0
        #     ]

        #     return row_index_groupe

        def reed_workbook(input_file):
            workbook = load_workbook(input_file)
            data_sheet = workbook.active
            # group_rows = get_group_rows(data_sheet)

            # vendor_name = None
            vendor_qs = None
            supplier_qs = None
            lot_auto = Lot.objects.get(name_shorts="шт")

            for index in range(3, data_sheet.max_row):
                row_level = data_sheet.row_dimensions[index].outline_level + 1
                item_value = data_sheet.cell(row=index, column=2).value

                # работа с стоками поставщиков
                if row_level == 1:
                    vendor_str = data_sheet.cell(row=index, column=1).value
                    vendor_str = vendor_str.replace(",", "").strip()
                    if (
                        vendor_str == "Товары"
                        or vendor_str == "Разное"
                        or vendor_str == ""
                        or vendor_str == "Оборудование (объекты основных средств)"
                    ):
                        vendor_str = "Hеизвестный"
                    elif vendor_str == "OptimusDrive":
                        vendor_str = "Optimus Drive"
                    elif vendor_str == "Delta":
                        vendor_str = "Delta Electronics"

                    slugish = translit.translify(vendor_str)
                    slugish = slugify(slugish)

                    try:
                        vendor_qs = Vendor.objects.get(slug=slugish)
                        supplier_qs = vendor_qs.supplier

                    except Vendor.DoesNotExist:
                        supplier_qs = Supplier.objects.get(slug="drugoj")
                        vat_catalog = Vat.objects.get(name=20)
                        currency_catalog = Currency.objects.get(words_code="RUB")

                        Vendor.objects.create(
                            name=vendor_str,
                            supplier=supplier_qs,
                            vat_catalog=vat_catalog,
                            currency_catalog=currency_catalog,
                        )

                # работа с товарами поставщиков
                else:
            
                    article_supplier = data_sheet.cell(row=index, column=1).value
                    all_fredom_remaining = data_sheet.cell(row=index, column=16).value
                
                    all_reserve_remaining = data_sheet.cell(row=index, column=15).value
                    if all_fredom_remaining:
                        int_stock_motrum = int(all_fredom_remaining)
                    else:
                        int_stock_motrum = 0  
                        
                    if all_reserve_remaining :    
                        int_stock_reserve_motrum = int(all_reserve_remaining)
                    else:
                        int_stock_reserve_motrum = 0
                    


                    if article_supplier != "" and article_supplier != None:
                        article_supplier = article_supplier.strip()
                        article_supplier = " ".join(article_supplier.split())
                        # товары находяться в окт с артикулом и производителем
                        try:
                            product = Product.objects.get(
                                vendor=vendor_qs, article_supplier=article_supplier
                            )
                            add_stok_motrum_old_article(product,lot_auto,int_stock_motrum,int_stock_reserve_motrum)
                        
                        # товары НЕ находяться в окт с артикулом и производителем
                        except Product.DoesNotExist:
                
                            error = "info_error"
                            location = "Загрузка фаилов остатков"

                            info = f"Загрузка фаилов остатков- такого товара нет в ОКТ. Добавьте вручную в окт товар {article_supplier}"
                            e = error_alert(error, location, info)
                            
                            # if vendor_qs.slug == "emas" or vendor_qs.slug == "tbloc":
                            #     try:
                            #         product = Product.objects.get(
                            #             supplier=supplier_qs, article_supplier=article_supplier
                            #         )
                            #         product.vendor = vendor_qs
                                    
                            #         product.save()
                            #         update_change_reason(product, "Автоматическое")
                            #         add_stok_motrum_old_article(product,lot_auto,int_stock_motrum,int_stock_reserve_motrum)
                            #     except Product.DoesNotExist:   
                            #         add_new_product_and_stock(supplier_qs,article_supplier,vendor_qs,lot_auto,int_stock_motrum,int_stock_reserve_motrum)
                                
                                    
                            # # порлностью новый товар
                            # else:
                            #     add_new_product_and_stock(supplier_qs,article_supplier,vendor_qs,lot_auto,int_stock_motrum,int_stock_reserve_motrum)

        def reed_workbook_after1c(input_file):
            
            lot_auto = Lot.objects.get(name_shorts="шт")
            workbook = load_workbook(input_file)
            data_sheet = workbook.active
            print(data_sheet.max_row)
            for index in range(2, data_sheet.max_row+1):
                
                article = data_sheet.cell(row=index, column=1).value
                if article != "" :
                    article = article.strip()
                    all_fredom_remaining = data_sheet.cell(row=index, column=4).value
                    print(all_fredom_remaining)
                    all_reserve_remaining = data_sheet.cell(row=index, column=5).value
                    print(all_reserve_remaining)
                    if all_fredom_remaining:
                        int_stock_motrum = int(all_fredom_remaining)
                    else:
                        int_stock_motrum = 0
                        
                    if all_reserve_remaining:
                        int_stock_reserve_motrum = int(all_reserve_remaining)
                    else:
                        int_stock_reserve_motrum = 0  
                    print("int_stock_reserve_motrum",int_stock_motrum,int_stock_reserve_motrum)
                    
                    try:
                        product = Product.objects.get(article=article)
                        print(333)
                        print("product",product)
                        add_stok_motrum_old_article(product,lot_auto,int_stock_motrum,int_stock_reserve_motrum)
                    
                    # товары НЕ находяться в окт с артикулом и производителем
                    except Product.DoesNotExist:
                        print("NOT PROD")
                        error = "info_error"
                        location = "Загрузка фаилов остатков"

                        info = f"Загрузка фаилов остатков- такого товара нет в ОКТ.Артикул мотрум {article}"
                        e = error_alert(error, location, info)    
                
               
                    
                    
                    
                    
                    
        # reed_workbook(path_storage_pnm)
        # reed_workbook(path_storage_motrum)
        reed_workbook_after1c(path_storage)
    except Exception as e:
        print(e)
        tr =  traceback.format_exc()
        error = "file_error"
        location = "Обработка фаилов остатков мотрум"

        info = f"ошибка при чтении фаила{e}{tr}"
        e = error_alert(error, location, info)
            

# вспомогательные функции
def add_new_product_and_stock(supplier_qs,article_supplier,vendor_qs,lot_auto,int_stock_motrum,int_stock_reserve_motrum):
    prod_new = Product(
        supplier=supplier_qs,
        name=article_supplier,
        vendor=vendor_qs,
        article_supplier=article_supplier,
        category_supplier_all=None,
        group_supplier=None,
        category_supplier=None,
    )
    prod_new.save()
    update_change_reason(prod_new, "Автоматическое")
    product_stock = Stock(
        prod=prod_new,
        lot=lot_auto,
        stock_motrum=int_stock_motrum,
        stock_motrum_reserve = int_stock_reserve_motrum,
    )
    product_stock.save()
    update_change_reason(product_stock, "Автоматическое")
    
    currency = Currency.objects.get(words_code="RUB")
    vat = Vat.objects.get(name=20)
    price =  Price(prod=prod_new,currency=currency,vat=vat,extra_price=True)
    price.save()
    update_change_reason(price, "Автоматическое")
    
def add_stok_motrum_old_article(product,lot_auto,int_stock_motrum,int_stock_reserve_motrum):
    product_stock = Stock.objects.get_or_create(
        prod=product,
        defaults={
            "lot": lot_auto,
            
        },
    )
    
    data_now =  datetime.datetime.today().date()
    print(type(data_now))

    
    
    stock = product_stock[0]
    print(stock.data_update_motrum )
    old_stock_data =  stock.data_update
    if stock.data_update_motrum == data_now:
        print(999999999)
        stock.stock_motrum = stock.stock_motrum + int_stock_motrum
        print(stock.stock_motrum_reserve )
        print(int_stock_reserve_motrum )
        stock.stock_motrum_reserve = stock.stock_motrum_reserve + int_stock_reserve_motrum
    else:
        stock.stock_motrum = int_stock_motrum
        stock.stock_motrum_reserve = int_stock_reserve_motrum
    
    stock._change_reason = "Автоматическое"
    stock.save()
    stock = Stock.objects.filter(prod=product).update(data_update=old_stock_data)

    # if old_stock_data:
    #     stock.data_update = old_stock_data
    # update_change_reason(stock, "Автоматическое")
    
